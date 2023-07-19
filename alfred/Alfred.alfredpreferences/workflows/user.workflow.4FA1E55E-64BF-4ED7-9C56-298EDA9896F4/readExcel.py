# -*- coding: utf-8 -*-
import sys
from openpyxl import load_workbook
reload(sys) # Python2.5 初始化后会删除 sys.setdefaultencoding 这个方法，我们需要重新载入
sys.setdefaultencoding('utf-8')
# sys.path.insert(0,'./')

# sys.path.insert(0,'./workflow')

# print (str(sys.path))
import workflow
# sys.setdefaultencoding('utf8') 

#print (str(sys.argv))
def read_excel(wf):
    excelName=sys.argv[1]
    zone=''
    key=''
    num1=-1
    index=2
    #print (len(sys.argv))
    if(len(sys.argv)>2):
        zone=sys.argv[2]
        if(len(sys.argv)>3):
            key=sys.argv[3]
            if(len(sys.argv)>4):
                num1=int(sys.argv[4])
    else:
        wf.add_item(title="请输入大区",subtitle="",arg='',
                 autocomplete="", valid=True, uid='')
        return
    
    # wf.add_item(title="1",subtitle="2")
    # 打开文件
    wb = load_workbook(excelName,data_only=True)
    # 获取所有sheet
    sheets = wb.worksheets
    sheet1=sheets[0]
    rowNum=sheet1.max_row
    # colNum=sheet1.max_column
    num=0
    #print (excelName,rowNum)
    for n in range(1,rowNum):
        id=sheet1.cell(n,1).value
        if(id==None):
            continue

        tzone=sheet1.cell(n,12).value
       
        if(tzone==None):
            continue

        if(str(tzone).find(str(zone))==0):
            server_type=sheet1.cell(n,11).value
       #print (n,id,tzone,server_type)
            if(server_type==None):
                continue
       
            title1=gbk2utf(server_type+str(id),0)
            arg1=gbk2utf(str(sheet1.cell(n,5).value),0)
        
            if(len(key)>0):
                #print(title1,arg1,server_type,id,num)
                if(str(server_type).find(str(key))==0):
                
                    #wf.add_item(title=str(server_type+" "+id), subtitle="", arg=str(sheet1.cel(n,5)._value),autocomplete="", valid=True, uid=str(n))
                    if(num1>0):
                        if(id==num1):
                            wf.add_item(title=str(title1), subtitle=str(tzone), arg=str(arg1), autocomplete="", valid=True, uid=str(n))
                    else:
                        wf.add_item(title=str(title1), subtitle=str(tzone), arg=str(arg1), autocomplete="", valid=True, uid=str(n))
                    num+=1
                    if(num>20):
                        break;
            else:
                wf.add_item(title=str(title1), subtitle=str(tzone), arg=str(arg1), autocomplete="", valid=True, uid=str(n)) 
       

    if(num==0):
        wf.add_item(title="无法找到大区或者服务器类型错误",subtitle="",arg='',
                 autocomplete="", valid=True, uid='')
    wf.send_feedback()

def gbk2utf(in_data , tag):
	if 1 == tag:
		return in_data.encode('gbk').decode('gbk')
	elif 0 == tag:
		return in_data.encode('gbk').decode('gbk').encode('utf8')

if __name__ == '__main__':
    wf = workflow.Workflow()
    wf.run(read_excel)